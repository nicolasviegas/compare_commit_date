# -*- coding: utf-8 -*-
"""
Created on Fri Mar 31 15:48:54 2023

@author: nviegas001
"""
import openpyxl
from openpyxl import load_workbook
import pandas as pd
#import os
#import shutil
import tkinter as tk
from tkinter import simpledialog
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

# PATH ES LA DIRECCION DONDE SE ENCUENTRA EL EXCEL
# PARENT DIR ES LA DIR DONDE SE VA A EJECUTAR EL SCRIPT Y DONDE SE VAN A CREAR LAS CARPETAS DE LOS CASOS
path = "/Users/nicolasviegas/Documents/full_screenshot.py/examples/excel.xlsx"
# parent_dir = "C:\\Users\\nviegas001\\python-scripts\\"
# download_dir = "C:\\Users\\nviegas001\\Downloads\\"
github_list = []
date_list = []

root = tk.Tk()
root.withdraw()
period = simpledialog.askstring(title="Compare commits",
                                prompt="Ingrese el periodo en el que se valida la fecha (I / U)")

root = tk.Tk()
root.withdraw()
pw = simpledialog.askstring(title="Compare commits",
                                prompt="Ingrese password")


def code_list_to_analyze():
    wb_obj = openpyxl.load_workbook(path)

    sheet_obj = wb_obj.active

    cell_obj = sheet_obj.cell(row=1, column=1)
    max_col = sheet_obj.max_column
    max_r = sheet_obj.max_row

    for i in range(2, max_r + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)

        link_github = cell_obj.value

        link = str(link_github)

        github_list.append(str(link))


def date_list_add(commit_date):
    individual_date = str(commit_date)
    date_list.append(str(individual_date))


def obtain_date_commit(chrome, link):
    chrome.get(link)
    commit_date = chrome.find_element(By.XPATH, "//relative-time[@class='no-wrap']").get_attribute("title")

    date_list_add(commit_date)


def open_list_links():
    chrome_options = webdriver.ChromeOptions()
    #chrome_options.add_argument('--headless') #No funciona en 2do plano
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_experimental_option('detach', True)

    chrome = webdriver.Chrome(executable_path=r"/Users/nicolasviegas/Documents/webdriver\chromedriver_mac64",
                              options=chrome_options)

    chrome.get(github_list[0])

    chrome.find_element(By.ID, "login_field").send_keys("nicolasviegas")
    chrome.find_element(By.ID, "password").send_keys(pw)
    chrome.find_element(By.NAME, 'commit').click()

    for i in github_list:
        if i != 'None':
            obtain_date_commit(chrome, i)

    chrome.quit()


def write_excel_file():
    wb_obj = openpyxl.load_workbook(path)

    sheet_obj = wb_obj.active

    if period == 'I' or period == 'i':
        #max_r = sheet_obj.max_row
        #for i in range(0, max_r-1):
        print("Entre por el if")
        workbook = openpyxl.load_workbook(path)
        worksheet = workbook.worksheets[0]
        worksheet.insert_cols(2)
        y = 2

        cell_title = worksheet.cell(row=1, column=2)
        cell_title.value = 'Fecha WT'

        for x in range(len(date_list)):
            cell_to_write = worksheet.cell(row=y, column=2)
            cell_to_write.value = date_list[x]
            print(date_list[x])
            y += 1

        workbook.save(path)

    elif period == 'U' or period == 'u':
        print("Entre por el elif")
        workbook = openpyxl.load_workbook(path)
        worksheet = workbook.worksheets[0]
        worksheet.insert_cols(3)
        y = 2

        cell_title = worksheet.cell(row=1, column=3)
        cell_title.value = 'Fecha Update'

        for x in range(len(date_list)):
            cell_to_write = worksheet.cell(row=y, column=3)
            cell_to_write.value = date_list[x]
            print(date_list[x])
            y += 1

        workbook.save(path)

    else:
        print("Ingrese una periodo valido ( I | U)")
        quit()



def request_validation():
    wb_obj = openpyxl.load_workbook(path)

    sheet_obj = wb_obj.active

    cell_obj = sheet_obj.cell(row=1, column=2)
    max_col = sheet_obj.max_column
    max_r = sheet_obj.max_row

    print(max_r)
    for i in range(2, max_r + 1):
        cell_obj = sheet_obj.cell(row=i, column=2)

        date_wt = cell_obj.value
        cell_obj_u = sheet_obj.cell(row=i, column=3)
        date_u = cell_obj_u.value
        if date_wt != date_u:
            print("Requiere validacion")
        else:
            print("No requiere validacion")



def main():

    code_list_to_analyze()

    open_list_links()

    write_excel_file()

    request_validation()


if __name__ == "__main__":
    main()
